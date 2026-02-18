#!/usr/bin/env python3
"""
Extract all data from NSW + QLD Excel files into harris_farm.db.

Source files:
  - _trading Team Meeting_Retail_NSW.xlsx (27 NSW stores, FY2017-FY2024)
  - _trading Team Meeting_QLD.xlsx (3 QLD stores, FY2021-FY2024)

Sheets extracted:
  - Ref → fiscal_calendar table
  - Data_ByMajGrp → sales table (wide-to-long melt)
  - Data_Customer → customers table (wide-to-long melt)

Excel layout (Data_ByMajGrp) — 0-based column indices:
  Row 15 (iter_rows 0-based from min_row=15): FY values at cols 11+
  Row 16: Metadata headers (cols 1-10) + week-of-year numbers (col 11+)
  Row 17+: Data rows

Uses iter_rows() for fast bulk reading instead of cell-by-cell access.
"""

import sqlite3
import sys
from pathlib import Path

import openpyxl
import pandas as pd

DATA_DIR = Path(__file__).parent.parent / "data"
DB_PATH = DATA_DIR / "harris_farm.db"

NSW_FILE = Path("/Users/angusharris/Downloads/Weekly Fresh Report/_trading Team Meeting_Retail_NSW.xlsx")
QLD_FILE = Path("/Users/angusharris/Downloads/Weekly Fresh Report/_trading Team Meeting_QLD.xlsx")


def safe_date(v):
    """Convert to YYYY-MM-DD string, or None if not a valid date."""
    if v is None or str(v).strip().lower() in ("", "--", "n/a"):
        return None
    try:
        return pd.Timestamp(v).strftime("%Y-%m-%d")
    except Exception:
        return None


def extract_ref(wb, label: str) -> pd.DataFrame:
    """Extract fiscal calendar from Ref sheet."""
    print(f"  [{label}] Reading Ref sheet...", flush=True)
    ws = wb["Ref"]

    rows = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        seq = row[1]  # col B (0-based index 1)
        if seq is None or str(seq) == "--":
            continue
        try:
            seq_int = int(seq)
        except (ValueError, TypeError):
            continue

        rows.append({
            "sequence": seq_int,
            "week_ending_date": safe_date(row[2]),      # col C
            "month_end": safe_date(row[4]),              # col E
            "prior_month_end": safe_date(row[5]),        # col F
            "retail_weeks_in_month": row[6],             # col G
            "financial_year": int(row[7]) if row[7] else None,   # col H
            "fin_week_of_month": int(row[8]) if row[8] else None, # col I
            "fin_week_of_year": int(row[9]) if row[9] else None,  # col J
            "prior_year_week_ending": safe_date(row[10] if len(row) > 10 else None),  # col K
            "current_fy_start": safe_date(row[11] if len(row) > 11 else None),        # col L
            "prior_fy_start": safe_date(row[12] if len(row) > 12 else None),          # col M
        })

    df = pd.DataFrame(rows)
    print(f"  [{label}] Ref: {len(df)} rows, FY {df['financial_year'].min()}-{df['financial_year'].max()}", flush=True)
    return df


def build_date_lookup(ref_df: pd.DataFrame) -> dict:
    """Build (fy, week_of_year) -> week_ending_date mapping."""
    lookup = {}
    for _, row in ref_df.iterrows():
        fy = row["financial_year"]
        wk = row["fin_week_of_year"]
        if pd.notna(fy) and pd.notna(wk):
            lookup[(int(fy), int(wk))] = row["week_ending_date"]
    return lookup


def extract_sales(wb, date_lookup: dict, label: str) -> pd.DataFrame:
    """Extract Data_ByMajGrp sheet using fast iter_rows bulk read."""
    print(f"  [{label}] Reading Data_ByMajGrp headers...", flush=True)
    ws = wb["Data_ByMajGrp"]

    # Read FY row (row 15) and week-num row (row 16) in bulk
    fy_row = None
    wk_row = None
    for row in ws.iter_rows(min_row=15, max_row=15, values_only=True):
        fy_row = list(row)
    for row in ws.iter_rows(min_row=16, max_row=16, values_only=True):
        wk_row = list(row)

    if not fy_row or not wk_row:
        print(f"  [{label}] ERROR: Could not read header rows")
        return pd.DataFrame()

    # Build week column mapping: list of (col_0based_index, week_ending_date)
    # Data columns start at index 11 (col L in Excel, col 12 in 1-based)
    week_map = []  # (0-based col index, week_ending_date)
    for i in range(11, len(fy_row)):
        fy_val = fy_row[i]
        wk_val = wk_row[i]
        if fy_val is None or str(fy_val) == "--":
            continue
        try:
            fy = int(fy_val)
            wk = int(wk_val)
        except (ValueError, TypeError):
            continue
        we_date = date_lookup.get((fy, wk))
        if we_date:
            week_map.append((i, we_date))

    print(f"  [{label}] Found {len(week_map)} valid week columns", flush=True)
    if not week_map:
        return pd.DataFrame()

    # Read ALL data rows at once using iter_rows (much faster than cell-by-cell)
    print(f"  [{label}] Reading data rows (bulk)...", flush=True)
    all_rows = []
    row_count = 0
    for row in ws.iter_rows(min_row=17, values_only=True):
        store = row[5]  # col F (0-based index 5) = Company Full Name
        if store is None or str(store) == "--":
            if row_count > 0:
                # Allow a few blank rows before giving up
                if all(r is None or str(r) == "--" for r in row[:11]):
                    break
            continue

        store_s = str(store)
        channel = str(row[6] or "")       # col G
        dept = str(row[7] or "")           # col H
        major_group = str(row[8] or "")    # col I
        is_promo = str(row[9] or "N")      # col J
        measure = str(row[10] or "")       # col K (Values)
        fy_lol = row[1]                    # col B
        rolling_13wk = row[2]              # col C
        fv_store = row[3]                  # col D

        fy_lol_int = int(fy_lol) if fy_lol is not None else None
        r13_int = int(rolling_13wk) if rolling_13wk is not None else None
        fv_int = int(fv_store) if fv_store is not None else None

        for col_idx, we_date in week_map:
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or str(val) == "--":
                continue
            try:
                val_float = float(val)
            except (ValueError, TypeError):
                continue
            if val_float == 0:
                continue

            all_rows.append((
                store_s, channel, dept, major_group, is_promo,
                measure, we_date, val_float,
                fy_lol_int, r13_int, fv_int,
            ))

        row_count += 1
        if row_count % 1000 == 0:
            print(f"  [{label}] Processed {row_count} rows ({len(all_rows):,} records)...", flush=True)

    cols = ["store", "channel", "department", "major_group", "is_promotion",
            "measure", "week_ending", "value", "fy_lol", "rolling_13wk_lol", "fv_store"]
    df = pd.DataFrame(all_rows, columns=cols)
    print(f"  [{label}] Sales: {row_count} data rows → {len(df):,} records", flush=True)
    if len(df) > 0:
        print(f"  [{label}]   {df['store'].nunique()} stores, {df['measure'].nunique()} measures", flush=True)
    return df


def extract_customers(wb, date_lookup: dict, label: str) -> pd.DataFrame:
    """Extract Data_Customer sheet using fast iter_rows bulk read."""
    print(f"  [{label}] Reading Data_Customer headers...", flush=True)
    ws = wb["Data_Customer"]

    # Read FY row (row 14) and header/week row (row 15)
    fy_row = None
    wk_row = None
    for row in ws.iter_rows(min_row=14, max_row=14, values_only=True):
        fy_row = list(row)
    for row in ws.iter_rows(min_row=15, max_row=15, values_only=True):
        wk_row = list(row)

    if not fy_row or not wk_row:
        print(f"  [{label}] ERROR: Could not read customer header rows")
        return pd.DataFrame()

    # Find where week data starts (first col with a valid FY int in row 14)
    week_map = []
    for i in range(len(fy_row)):
        fy_val = fy_row[i]
        if fy_val is None or str(fy_val) == "--":
            continue
        try:
            fy = int(fy_val)
        except (ValueError, TypeError):
            continue
        wk_val = wk_row[i] if i < len(wk_row) else None
        if wk_val is None:
            continue
        try:
            wk = int(wk_val)
        except (ValueError, TypeError):
            continue
        we_date = date_lookup.get((fy, wk))
        if we_date:
            week_map.append((i, we_date))

    print(f"  [{label}] Found {len(week_map)} valid customer week columns", flush=True)
    if not week_map:
        return pd.DataFrame()

    # Read data rows (start at row 16)
    # Customer metadata: col 6=store (0-based), col 7=channel, col 8=measure
    all_rows = []
    row_count = 0
    for row in ws.iter_rows(min_row=16, values_only=True):
        store = row[6] if len(row) > 6 else None   # Company Full Name
        if store is None or str(store) == "--":
            if row_count > 0:
                if all((r is None or str(r) == "--") for r in row[:9] if r is not None):
                    break
            continue

        store_s = str(store)
        channel = str(row[7] or "") if len(row) > 7 else ""
        measure = str(row[8] or "") if len(row) > 8 else ""

        for col_idx, we_date in week_map:
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or str(val) == "--":
                continue
            try:
                val_float = float(val)
            except (ValueError, TypeError):
                continue
            if val_float == 0:
                continue

            all_rows.append((store_s, channel, measure, we_date, val_float))

        row_count += 1

    cols = ["store", "channel", "measure", "week_ending", "value"]
    df = pd.DataFrame(all_rows, columns=cols)
    print(f"  [{label}] Customers: {row_count} data rows → {len(df):,} records", flush=True)
    return df


def extract_stores(sales_df: pd.DataFrame) -> pd.DataFrame:
    """Build stores table from sales data."""
    stores = sales_df["store"].unique()
    rows = []
    for s in sorted(stores):
        parts = s.split(" - ", 1)
        num = parts[0].strip() if len(parts) > 1 else ""
        name = parts[1].strip() if len(parts) > 1 else s
        try:
            num_int = int(num)
            # Only these specific stores are QLD
            state = "QLD" if num_int in (70, 74, 75) else "NSW"
        except ValueError:
            state = "NSW"
        rows.append({
            "store_number": num,
            "store_name": name,
            "store_company": "Harris Farm Markets",
            "state": state,
            "store_type": "Retail",
            "fv_store": 1,
            "like_for_like": 1,
        })
    return pd.DataFrame(rows)


def extract_departments(sales_df: pd.DataFrame) -> pd.DataFrame:
    """Build departments table from sales data."""
    pairs = sales_df[["department", "major_group"]].drop_duplicates()
    rows = []
    for _, row in pairs.iterrows():
        dept = row["department"]
        mg = row["major_group"]
        dept_parts = dept.split(" - ", 1)
        mg_parts = mg.split(" - ", 1)
        rows.append({
            "department_code": dept_parts[0].strip() if len(dept_parts) > 1 else "",
            "department_name": dept_parts[1].strip() if len(dept_parts) > 1 else dept,
            "major_group_code": mg_parts[0].strip() if len(mg_parts) > 1 else "",
            "major_group_name": mg_parts[1].strip() if len(mg_parts) > 1 else mg,
        })
    return pd.DataFrame(rows).drop_duplicates().sort_values(["department_code", "major_group_code"])


def main():
    print("=" * 60, flush=True)
    print("Harris Farm Hub — Full Data Extraction", flush=True)
    print("=" * 60, flush=True)

    for f in [NSW_FILE, QLD_FILE]:
        if not f.exists():
            print(f"ERROR: {f} not found")
            sys.exit(1)

    DATA_DIR.mkdir(exist_ok=True)

    # Backup existing DB
    if DB_PATH.exists():
        backup = DB_PATH.with_suffix(".db.bak")
        import shutil
        shutil.copy2(DB_PATH, backup)
        print(f"Backed up existing DB to {backup.name}", flush=True)

    # --- NSW ---
    print("\n--- NSW ---", flush=True)
    print("Loading workbook...", flush=True)
    wb_nsw = openpyxl.load_workbook(str(NSW_FILE), read_only=True, data_only=True)

    ref_nsw = extract_ref(wb_nsw, "NSW")
    date_lookup_nsw = build_date_lookup(ref_nsw)
    sales_nsw = extract_sales(wb_nsw, date_lookup_nsw, "NSW")
    customers_nsw = extract_customers(wb_nsw, date_lookup_nsw, "NSW")
    wb_nsw.close()

    # --- QLD ---
    print("\n--- QLD ---", flush=True)
    print("Loading workbook...", flush=True)
    wb_qld = openpyxl.load_workbook(str(QLD_FILE), read_only=True, data_only=True)

    ref_qld = extract_ref(wb_qld, "QLD")
    date_lookup_qld = build_date_lookup(ref_qld)
    sales_qld = extract_sales(wb_qld, date_lookup_qld, "QLD")
    customers_qld = extract_customers(wb_qld, date_lookup_qld, "QLD")
    wb_qld.close()

    # --- Merge ---
    print("\n--- Merging ---", flush=True)
    ref_all = pd.concat([ref_nsw, ref_qld]).drop_duplicates(
        subset=["financial_year", "fin_week_of_year"], keep="first"
    ).sort_values("sequence").reset_index(drop=True)
    print(f"Fiscal calendar: {len(ref_all)} rows", flush=True)

    sales_all = pd.concat([sales_nsw, sales_qld], ignore_index=True)
    print(f"Sales: {len(sales_all):,} total records", flush=True)
    print(f"  Stores: {sales_all['store'].nunique()}", flush=True)
    print(f"  Measures: {sorted(sales_all['measure'].unique())}", flush=True)

    customers_all = pd.concat([customers_nsw, customers_qld], ignore_index=True)
    print(f"Customers: {len(customers_all):,} total records", flush=True)

    stores_df = extract_stores(sales_all)
    print(f"Stores: {len(stores_df)} stores", flush=True)

    depts_df = extract_departments(sales_all)
    print(f"Departments: {len(depts_df)} dept/major-group combos", flush=True)

    # --- Write to SQLite ---
    print("\n--- Writing to SQLite ---", flush=True)
    conn = sqlite3.connect(str(DB_PATH))

    conn.execute("DROP TABLE IF EXISTS fiscal_calendar")
    conn.execute("DROP TABLE IF EXISTS sales")
    conn.execute("DROP TABLE IF EXISTS customers")
    conn.execute("DROP TABLE IF EXISTS stores")
    conn.execute("DROP TABLE IF EXISTS departments")

    ref_all.to_sql("fiscal_calendar", conn, index=False)
    print(f"  fiscal_calendar: {len(ref_all)} rows written", flush=True)

    sales_all.to_sql("sales", conn, index=False, chunksize=50000)
    print(f"  sales: {len(sales_all):,} rows written", flush=True)

    customers_all.to_sql("customers", conn, index=False)
    print(f"  customers: {len(customers_all):,} rows written", flush=True)

    stores_df.to_sql("stores", conn, index=False)
    print(f"  stores: {len(stores_df)} rows written", flush=True)

    depts_df.to_sql("departments", conn, index=False)
    print(f"  departments: {len(depts_df)} rows written", flush=True)

    # Keep existing tables
    existing = [r[0] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table'"
    ).fetchall()]
    for t in ["market_share", "product_lines"]:
        if t in existing:
            count = conn.execute(f"SELECT COUNT(*) FROM [{t}]").fetchone()[0]
            print(f"  {t}: {count} rows (kept as-is)", flush=True)

    # Create indexes
    print("\n  Creating indexes...", flush=True)
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_store ON sales(store)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_week ON sales(week_ending)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_measure ON sales(measure)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sales_dept ON sales(department)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_customers_store ON customers(store)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_customers_week ON customers(week_ending)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_fc_fy_wk ON fiscal_calendar(financial_year, fin_week_of_year)")
    print("  Done.", flush=True)

    conn.commit()
    conn.close()

    # --- Summary ---
    print("\n" + "=" * 60, flush=True)
    print("EXTRACTION COMPLETE", flush=True)
    print("=" * 60, flush=True)
    db_size = DB_PATH.stat().st_size / (1024 * 1024)
    print(f"Database: {DB_PATH} ({db_size:.1f} MB)", flush=True)
    print(f"Sales: {len(sales_all):,} records ({sales_all['store'].nunique()} stores)", flush=True)
    print(f"Customers: {len(customers_all):,} records", flush=True)
    print(f"Calendar: {len(ref_all)} weeks", flush=True)
    print(f"Stores: {len(stores_df)}, Dept combos: {len(depts_df)}", flush=True)


if __name__ == "__main__":
    main()
